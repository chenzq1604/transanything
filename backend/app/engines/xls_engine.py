"""
XLS/XLSX转换引擎
使用openpyxl读取Excel文件，每个Sheet转为一个Markdown表格，处理合并单元格
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.engines.base import BaseEngine


class XlsEngine(BaseEngine):
    """Excel文件转换引擎，将每个Sheet转为Markdown表格"""

    @property
    def supported_extensions(self) -> list[str]:
        """返回Excel引擎支持的扩展名"""
        return [".xls", ".xlsx"]

    async def convert(self, file_path: str) -> str:
        """
        将Excel文件转换为Markdown格式

        Args:
            file_path: Excel文件路径

        Returns:
            转换后的Markdown文本，每个Sheet一个表格
        """
        wb = load_workbook(file_path, read_only=True, data_only=True)
        markdown_parts: list[str] = []

        try:
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # 添加Sheet标题
                markdown_parts.append(f"## {sheet_name}\n\n")

                # 转换Sheet为Markdown表格
                md_table = self._sheet_to_markdown(sheet)
                if md_table:
                    markdown_parts.append(md_table + "\n\n")
        finally:
            wb.close()

        return "".join(markdown_parts).strip()

    def _sheet_to_markdown(self, sheet: Worksheet) -> str:
        """
        将Excel Sheet转换为Markdown表格

        Args:
            sheet: openpyxl的Worksheet对象

        Returns:
            Markdown格式的表格字符串
        """
        # 获取合并单元格映射
        merged_map = self._build_merged_map(sheet)

        # 获取有效数据范围
        max_row = sheet.max_row
        max_col = sheet.max_column

        if not max_row or not max_col:
            return ""

        lines: list[str] = []

        for row_idx in range(1, max_row + 1):
            row_data: list[str] = []
            for col_idx in range(1, max_col + 1):
                cell_coord = f"{get_column_letter(col_idx)}{row_idx}"

                # 检查是否在合并单元格中
                if cell_coord in merged_map:
                    value = merged_map[cell_coord]
                else:
                    try:
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        value = cell.value
                    except Exception:
                        value = None

                # 格式化单元格值
                if value is None:
                    row_data.append("")
                else:
                    row_data.append(str(value).strip().replace("\n", " "))

            lines.append("| " + " | ".join(row_data) + " |")

            # 第一行后添加分隔线（表头）
            if row_idx == 1:
                lines.append("| " + " | ".join(["---"] * max_col) + " |")

        return "\n".join(lines)

    def _build_merged_map(self, sheet: Worksheet) -> dict[str, str]:
        """
        构建合并单元格映射表

        将合并区域中的所有单元格映射到左上角的值，
        非左上角的单元格标记为空字符串（避免重复显示）

        Args:
            sheet: openpyxl的Worksheet对象

        Returns:
            单元格坐标到值的映射字典
        """
        merged_map: dict[str, str] = {}

        for merged_range in sheet.merged_cells.ranges:
            min_row = merged_range.min_row
            min_col = merged_range.min_col
            max_row = merged_range.max_row
            max_col = merged_range.max_col

            # 获取合并区域左上角的值
            try:
                top_left_value = sheet.cell(row=min_row, column=min_col).value
            except Exception:
                top_left_value = None

            value_str = str(top_left_value).strip() if top_left_value is not None else ""

            # 左上角单元格保留值
            top_left_coord = f"{get_column_letter(min_col)}{min_row}"
            merged_map[top_left_coord] = value_str

            # 其他单元格设为空（避免重复）
            for row_idx in range(min_row, max_row + 1):
                for col_idx in range(min_col, max_col + 1):
                    if row_idx == min_row and col_idx == min_col:
                        continue
                    coord = f"{get_column_letter(col_idx)}{row_idx}"
                    merged_map[coord] = ""

        return merged_map
