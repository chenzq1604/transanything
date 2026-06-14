"""
文件管理API
提供文件列表、详情、下载、删除、预览、内容获取和多引擎结果管理功能
"""

import logging
import re
from pathlib import Path

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, PlainTextResponse

from app.config import settings
from app.models.schemas import FileDetailResponse, FileListResponse, FileUpload, SaveMarkdownRequest
from app.services.convert_service import get_convert_service


def _fix_image_paths(markdown: str, file_id: str) -> str:
    """
    修正markdown中的图片路径，将本地相对路径转换为API路径
    兼容旧数据中HTML <img>标签使用本地路径(如outputs\\file_id\\images\\xxx.jpg)的情况
    """
    # 修正HTML <img src="outputs/..."> 或 <img src="outputs\\...">
    def fix_html_img(match):
        full_tag = match.group(0)
        src = match.group(1)
        # 已经是API路径的不处理
        if src.startswith("/api/"):
            return full_tag
        img_name = Path(src).name
        new_src = f"/api/files/{file_id}/images/{img_name}"
        return full_tag.replace(src, new_src)

    result = re.sub(r'<img\s+[^>]*src=["\']([^"\']+)["\']', fix_html_img, markdown)

    # 修正标准markdown图片 ![alt](outputs/...)
    def fix_md_img(match):
        alt_text = match.group(1) or ""
        img_path = match.group(2)
        if img_path.startswith("/api/"):
            return match.group(0)
        img_name = Path(img_path).name
        return f"![{alt_text}](/api/files/{file_id}/images/{img_name})"

    result = re.sub(r'!\[([^\]]*)\]\(([^)]+)\)', fix_md_img, result)

    return result


logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["文件管理"])


@router.get("/files", response_model=FileListResponse, summary="获取文件列表")
async def list_files() -> FileListResponse:
    """
    获取所有已上传文件的信息列表

    Returns:
        包含文件列表和总数的响应
    """
    service = get_convert_service()
    files = await service.list_files()
    return FileListResponse(files=files, total=len(files))


@router.get("/files/{file_id}", response_model=FileDetailResponse, summary="获取文件详情")
async def get_file_detail(file_id: str) -> FileDetailResponse:
    """
    获取指定文件的详细信息，包括转换后的Markdown内容和所有引擎结果

    兼容旧接口：markdown_content字段仍然返回默认引擎的结果。
    新增：engine_results字段返回所有引擎的结果列表。

    Args:
        file_id: 文件ID

    Returns:
        文件详情，包含基本信息、Markdown内容和引擎结果列表

    Raises:
        HTTPException: 文件不存在
    """
    service = get_convert_service()

    file_info = await service.get_file(file_id)
    if file_info is None:
        raise HTTPException(status_code=404, detail=f"文件不存在: {file_id}")

    markdown_content = await service.get_markdown_content(file_id)

    # 实时修正markdown中的图片路径（兼容旧数据中HTML img标签使用本地路径的情况）
    if markdown_content and file_id:
        markdown_content = _fix_image_paths(markdown_content, file_id)

    # 获取所有引擎结果
    engine_results = await service.get_engine_results(file_id)

    # 修正各引擎结果中的图片路径
    for er in engine_results:
        if er.markdown_content and file_id:
            er.markdown_content = _fix_image_paths(er.markdown_content, file_id)

    return FileDetailResponse(
        file_info=file_info,
        markdown_content=markdown_content,
        engine_results=engine_results,
    )


@router.get("/files/{file_id}/engines", summary="获取文件所有引擎结果列表")
async def get_engine_results(file_id: str) -> dict:
    """
    获取文件所有引擎的转换结果列表

    返回每个引擎的转换状态、内容摘要、耗时等信息。

    Args:
        file_id: 文件ID

    Returns:
        包含engine_results列表的响应

    Raises:
        HTTPException: 文件不存在
    """
    service = get_convert_service()

    file_info = await service.get_file(file_id)
    if file_info is None:
        raise HTTPException(status_code=404, detail=f"文件不存在: {file_id}")

    engine_results = await service.get_engine_results(file_id)

    # 修正各引擎结果中的图片路径
    for er in engine_results:
        if er.markdown_content and file_id:
            er.markdown_content = _fix_image_paths(er.markdown_content, file_id)

    return {
        "file_id": file_id,
        "engine_results": [er.model_dump() for er in engine_results],
    }


@router.put("/files/{file_id}/engines/{engine_type}/save", summary="保存指定引擎的Markdown")
async def save_engine_markdown(file_id: str, engine_type: str, request: SaveMarkdownRequest) -> dict:
    """
    保存指定引擎的用户编辑后Markdown内容

    同时更新该引擎的结果和默认的markdown_content（兼容旧接口）。

    Args:
        file_id: 文件ID
        engine_type: 引擎类型标识
        request: 包含编辑后Markdown内容的请求体

    Returns:
        保存结果信息

    Raises:
        HTTPException: 文件不存在或保存失败
    """
    service = get_convert_service()

    # 检查文件是否存在
    file_info = await service.get_file(file_id)
    if file_info is None:
        raise HTTPException(status_code=404, detail=f"文件不存在: {file_id}")

    try:
        output_path = await service.save_engine_markdown(file_id, engine_type, request.content)
        return {"message": "保存成功", "file_id": file_id, "engine_type": engine_type, "output_path": output_path}
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"保存引擎Markdown失败: {file_id}/{engine_type}, 错误: {str(e)}")
        raise HTTPException(status_code=500, detail=f"保存失败: {str(e)}")


@router.get("/files/{file_id}/download", summary="下载Markdown文件")
async def download_markdown(file_id: str) -> FileResponse:
    """
    下载指定文件转换后的Markdown文件

    Args:
        file_id: 文件ID

    Returns:
        Markdown文件下载响应

    Raises:
        HTTPException: 文件不存在或Markdown内容未生成
    """
    service = get_convert_service()

    # 检查文件是否存在
    file_info = await service.get_file(file_id)
    if file_info is None:
        raise HTTPException(status_code=404, detail=f"文件不存在: {file_id}")

    # 检查Markdown文件是否存在
    output_path = Path(settings.OUTPUT_DIR) / f"{file_id}.md"
    if not output_path.exists():
        raise HTTPException(status_code=404, detail="Markdown文件尚未生成，请先执行转换")

    # 使用原始文件名作为下载文件名
    original_name = Path(file_info.filename).stem
    download_filename = f"{original_name}.md"

    return FileResponse(
        path=str(output_path),
        filename=download_filename,
        media_type="text/markdown",
    )


@router.delete("/files/{file_id}", summary="删除文件")
async def delete_file(file_id: str) -> dict:
    """
    删除指定文件及其所有相关资源（原始文件、Markdown输出、引擎结果）

    Args:
        file_id: 文件ID

    Returns:
        删除结果信息

    Raises:
        HTTPException: 文件不存在或删除失败
    """
    service = get_convert_service()

    success = await service.delete_file(file_id)
    if not success:
        raise HTTPException(status_code=404, detail=f"文件不存在: {file_id}")

    return {"message": "删除成功", "file_id": file_id}


@router.get("/files/{file_id}/preview", summary="获取原始文件预览")
async def preview_file(file_id: str) -> FileResponse:
    """
    获取原始文件用于前端预览

    支持PDF、图片等可直接预览的文件类型

    Args:
        file_id: 文件ID

    Returns:
        原始文件响应

    Raises:
        HTTPException: 文件不存在
    """
    service = get_convert_service()

    file_path = await service.get_upload_file_path(file_id)
    if file_path is None:
        raise HTTPException(status_code=404, detail=f"文件不存在: {file_id}")

    # 根据文件类型确定MIME类型
    mime_types = {
        ".pdf": "application/pdf",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        ".txt": "text/plain; charset=utf-8",
        ".md": "text/markdown; charset=utf-8",
        ".csv": "text/csv; charset=utf-8",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
    }

    ext = file_path.suffix.lower()
    media_type = mime_types.get(ext, "application/octet-stream")

    return FileResponse(
        path=str(file_path),
        media_type=media_type,
        headers={"Content-Disposition": f"inline; filename=\"{file_path.name}\""},
    )


@router.get("/files/{file_id}/text-content", summary="获取文本文件内容")
async def get_text_content(file_id: str) -> PlainTextResponse:
    """
    获取文本类型文件的文本内容（用于前端TxtViewer展示）

    Args:
        file_id: 文件ID

    Returns:
        文本内容

    Raises:
        HTTPException: 文件不存在或不是文本文件
    """
    service = get_convert_service()

    file_path = await service.get_upload_file_path(file_id)
    if file_path is None:
        raise HTTPException(status_code=404, detail=f"文件不存在: {file_id}")

    # 尝试用多种编码读取文件
    text_extensions = {".txt", ".md", ".csv", ".log", ".json", ".xml", ".yaml", ".yml"}
    ext = file_path.suffix.lower()

    if ext not in text_extensions:
        # 非文本文件，尝试直接读取
        try:
            return PlainTextResponse(file_path.read_text(encoding="utf-8"))
        except (UnicodeDecodeError, PermissionError):
            raise HTTPException(status_code=400, detail="该文件不是文本格式，无法以文本方式展示")

    # 尝试不同编码读取
    for encoding in ["utf-8", "gbk", "gb2312", "latin-1"]:
        try:
            content = file_path.read_text(encoding=encoding)
            return PlainTextResponse(content)
        except (UnicodeDecodeError, PermissionError):
            continue

    raise HTTPException(status_code=500, detail="无法解码文件内容")


@router.get("/files/{file_id}/excel-data", summary="获取Excel文件结构化数据")
async def get_excel_data(file_id: str) -> dict:
    """
    获取Excel文件的结构化数据（用于前端XlsViewer展示）

    Args:
        file_id: 文件ID

    Returns:
        Excel各工作表的数据结构

    Raises:
        HTTPException: 文件不存在或不是Excel文件
    """
    import openpyxl

    service = get_convert_service()

    file_path = await service.get_upload_file_path(file_id)
    if file_path is None:
        raise HTTPException(status_code=404, detail=f"文件不存在: {file_id}")

    ext = file_path.suffix.lower()
    if ext not in {".xls", ".xlsx"}:
        raise HTTPException(status_code=400, detail="该文件不是Excel格式")

    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        sheets = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data = []
            headers = []

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = [str(cell) if cell is not None else "" for cell in row]
                else:
                    row_list = [str(cell) if cell is not None else "" for cell in row]
                    rows_data.append(row_list)

            # 如果没有数据，至少提供空表头
            if not headers and not rows_data:
                headers = ["(空)"]

            sheets[sheet_name] = {
                "headers": headers,
                "rows": rows_data,
            }

        wb.close()
        return {"sheets": sheets}

    except Exception as e:
        logger.error(f"读取Excel数据失败: {file_id}, 错误: {str(e)}")
        raise HTTPException(status_code=500, detail=f"读取Excel数据失败: {str(e)}")


@router.get("/files/{file_id}/images/{image_name}", summary="获取PDF提取的嵌入图片")
async def get_embedded_image(file_id: str, image_name: str) -> FileResponse:
    """
    获取PDF转换时提取的嵌入图片

    用于Markdown中的图片引用，如 ![图](/api/files/{id}/images/page1_img1.png)

    Args:
        file_id: 文件ID
        image_name: 图片文件名

    Returns:
        图片文件响应

    Raises:
        HTTPException: 图片不存在
    """
    image_path = Path(settings.OUTPUT_DIR) / file_id / "images" / image_name

    if not image_path.exists():
        raise HTTPException(status_code=404, detail=f"图片不存在: {image_name}")

    # 安全检查：防止路径遍历
    if ".." in image_name or "/" in image_name or "\\" in image_name:
        raise HTTPException(status_code=400, detail="非法图片名称")

    # 根据扩展名确定Content-Type
    content_types = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".gif": "image/gif",
        ".bmp": "image/bmp",
        ".webp": "image/webp",
    }
    ext = image_path.suffix.lower()
    media_type = content_types.get(ext, "application/octet-stream")

    return FileResponse(
        path=str(image_path),
        media_type=media_type,
        headers={"Cache-Control": "public, max-age=86400"},
    )
